from aiogram import Router, types, F, Bot
from aiogram.fsm.context import FSMContext
from aiogram.fsm.state import State, StatesGroup
from aiogram.types import Message, CallbackQuery, InlineKeyboardMarkup, InlineKeyboardButton, FSInputFile
from sqlalchemy.ext.asyncio import AsyncSession
from aiogram.exceptions import TelegramBadRequest
from skyde_bot.app.services.user_services import UserService
from skyde_bot.app.utils.utils import delete_previous_message, set_last_message_id
from skyde_bot.app.keyboards.inline import main_menu_keyboard
from sqlalchemy import select
from skyde_bot.app.database.models import User
import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill
from datetime import datetime
import os

# --- КОНФИГУРАЦИЯ ---
router = Router()


class DevState(StatesGroup):
    """Состояния для меню разработчика."""
    waiting_for_g_amount = State()  # Ожидание ввода суммы G-монет

    # Новые состояния для начисления баллов пользователю
    waiting_for_user_identifier = State()  # Ожидание username/user_id
    waiting_for_points_amount = State()  # Ожидание количества баллов
    waiting_for_points_comment = State()  # Ожидание комментария


# !!! ВАЖНО: Замените это на ваш фактический Telegram ID !!!
DEVELOPER_ID = 729292013


# ---------------------

# --- КЛАВИАТУРЫ ---

def dev_menu_keyboard() -> InlineKeyboardMarkup:
    """Клавиатура для меню разработчика."""
    return InlineKeyboardMarkup(inline_keyboard=[
        [InlineKeyboardButton(text="💰 Начислить G-баллы себе", callback_data="dev_add_g_balance")],
        [InlineKeyboardButton(text="🎁 Начислить баллы пользователю", callback_data="dev_add_user_balance")],
        [InlineKeyboardButton(text="📊 Экспорт базы пользователей", callback_data="dev_export_users")],
        [InlineKeyboardButton(text="💥 Сбросить FSM-состояния", callback_data="dev_clear_fsm")],
        [InlineKeyboardButton(text="⬅️ В главное меню", callback_data="main_menu_return")],
    ])


def dev_cancel_keyboard() -> InlineKeyboardMarkup:
    """Клавиатура отмены в режиме ввода."""
    return InlineKeyboardMarkup(inline_keyboard=[
        [InlineKeyboardButton(text="❌ Отмена", callback_data="dev_menu_return")]
    ])


def dev_confirm_keyboard() -> InlineKeyboardMarkup:
    """Клавиатура подтверждения действия."""
    return InlineKeyboardMarkup(inline_keyboard=[
        [
            InlineKeyboardButton(text="✅ Подтвердить", callback_data="dev_confirm_add_points"),
            InlineKeyboardButton(text="❌ Отмена", callback_data="dev_menu_return")
        ]
    ])


# --- ВСПОМОГАТЕЛЬНЫЙ ФУНКЦИОНАЛ ---

async def show_dev_menu(message: Message | CallbackQuery, state: FSMContext, bot: Bot):
    """Отправляет меню разработчика, удаляя предыдущее сообщение."""
    await state.clear()

    # Удаляем предыдущее сообщение, чтобы не было дубликатов
    await delete_previous_message(state, message.chat.id, bot)

    new_message = await message.answer(
        "🛠 <b>МЕНЮ АДМИНИСТРАТОРА</b> 🛠\n\n"
        "Выберите действие для управления ботом:",
        reply_markup=dev_menu_keyboard(),
        parse_mode='HTML'
    )
    # Сохраняем ID нового меню
    await set_last_message_id(state, new_message.message_id)


# --- ХЕНДЛЕРЫ ДЛЯ НАЧИСЛЕНИЯ БАЛЛОВ СЕБЕ ---

# 1. Запуск меню разработчика (команда /dev)
@router.message(F.text == "/dev", F.from_user.id == DEVELOPER_ID)
async def start_dev_menu_command(message: Message, state: FSMContext, bot: Bot):
    """Обрабатывает команду /dev."""
    await message.delete()  # Удаляем команду /dev
    # Вызываем вспомогательную функцию для отправки меню
    await show_dev_menu(message, state, bot)


# 2. Начать начисление G-баллов СЕБЕ (callback)
@router.callback_query(F.data == "dev_add_g_balance")
async def start_add_balance(callback: CallbackQuery, state: FSMContext, bot: Bot):
    """Переводит в состояние ожидания суммы для начисления."""
    await callback.answer()

    # Удаляем меню разработчика
    await delete_previous_message(state, callback.message.chat.id, bot)

    await state.set_state(DevState.waiting_for_g_amount)

    new_message = await callback.message.answer(
        "💰 <b>Начисление баллов себе</b>\n\n"
        "Введите сумму G-баллов (целое число), которую хотите начислить:",
        reply_markup=dev_cancel_keyboard(),
        parse_mode='HTML'
    )
    await set_last_message_id(state, new_message.message_id)

# 3. Обработка ввода суммы и начисление СЕБЕ
@router.message(DevState.waiting_for_g_amount)
async def process_add_balance(message: Message, state: FSMContext, session: AsyncSession, bot: Bot):
    """Проверяет ввод и начисляет G-баллы администратору."""

    # 1. Удаляем предыдущее сообщение бота (запрос суммы)
    await delete_previous_message(state, message.chat.id, bot)
    # 2. Удаляем сообщение пользователя (введенная сумма)
    await message.delete()

    try:
        amount = int(message.text.strip())
        if amount <= 0:
            raise ValueError

        user_service = UserService(session)

        new_balance = await user_service.update_balance(
            telegram_id=message.from_user.id,
            amount=amount
        )

        await state.clear()

        # 3. Отправляем подтверждение
        await message.answer(
            f"✅ <b>Успех!</b>\n\n"
            f"Начислено <b>{amount}</b> G-баллов.\n"
            f"💼 Новый баланс: <b>{new_balance}</b> G.",
            parse_mode='HTML'
        )

        # 4. Возвращаемся в меню разработчика (чистая отправка)
        await show_dev_menu(message, state, bot)

    except ValueError:
        new_error_message = await message.answer(
            "❌ Ошибка ввода. Введите корректное целое положительное число:",
            reply_markup=dev_cancel_keyboard(),
            parse_mode='HTML'
        )
        await set_last_message_id(state, new_error_message.message_id)

    except Exception as e:
        await state.clear()
        # Показываем чистую ошибку и возвращаем в меню
        await message.answer(f"❌ Непредвиденная ошибка: {e}", reply_markup=dev_menu_keyboard())
        await show_dev_menu(message, state, bot)

# --- НОВЫЕ ХЕНДЛЕРЫ: НАЧИСЛЕНИЕ БАЛЛОВ ПОЛЬЗОВАТЕЛЮ ---

# 4. Начало процесса начисления баллов пользователю
@router.callback_query(F.data == "dev_add_user_balance")
async def start_add_user_balance(callback: CallbackQuery, state: FSMContext, bot: Bot):
    """Начало процесса начисления баллов пользователю."""
    await callback.answer()

    await delete_previous_message(state, callback.message.chat.id, bot)

    await state.set_state(DevState.waiting_for_user_identifier)

    new_message = await callback.message.answer(
        "🎁 <b>Начисление баллов пользователю</b>\n\n"
        "Введите <b>@username</b> или <b>Telegram ID</b> пользователя:\n\n"
        "💡 Примеры:\n"
        "• <code>@username</code>\n"
        "• <code>123456789</code>",
        reply_markup=dev_cancel_keyboard(),
        parse_mode='HTML'
    )
    await set_last_message_id(state, new_message.message_id)

# 5. Обработка ввода username/user_id
@router.message(DevState.waiting_for_user_identifier)
async def process_user_identifier(message: Message, state: FSMContext, session: AsyncSession, bot: Bot):
    """Поиск пользователя по username или ID."""

    await delete_previous_message(state, message.chat.id, bot)
    await message.delete()

    identifier = message.text.strip()

    user_service = UserService(session)
    user = None

    # Пытаемся найти по username
    if identifier.startswith('@'):
        username = identifier.lstrip('@')
        result = await session.execute(
            select(User).where(User.username == username)
        )
        user = result.scalar_one_or_none()
    # Пытаемся найти по Telegram ID
    elif identifier.isdigit():
        telegram_id = int(identifier)
        user = await user_service.get_user_by_telegram_id(telegram_id)

    if not user:
        error_msg = await message.answer(
            "❌ <b>Пользователь не найден</b>\n\n"
            "Проверьте правильность ввода и попробуйте снова:\n"
            "• @username\n"
            "• Telegram ID",
            reply_markup=dev_cancel_keyboard(),
            parse_mode='HTML'
        )
        await set_last_message_id(state, error_msg.message_id)
        return

    # Сохраняем данные пользователя
    await state.update_data(
        target_user_id=user.telegram_id,
        target_user_name=user.full_name,
        target_username=user.username
    )

    await state.set_state(DevState.waiting_for_points_amount)

    amount_msg = await message.answer(
        f"✅ <b>Пользователь найден:</b>\n\n"
        f"👤 Имя: <b>{user.full_name}</b>\n"
        f"🆔 Username: @{user.username or 'не указан'}\n"
        f"💼 Баланс: <b>{user.balance_g}</b> G\n\n"
        f"💰 Введите количество баллов для начисления:",
        reply_markup=dev_cancel_keyboard(),
        parse_mode='HTML'
    )
    await set_last_message_id(state, amount_msg.message_id)

# 6. Обработка ввода количества баллов
@router.message(DevState.waiting_for_points_amount)
async def process_points_amount(message: Message, state: FSMContext, bot: Bot):
    """Обработка введенного количества баллов."""

    await delete_previous_message(state, message.chat.id, bot)
    await message.delete()

    try:
        amount = int(message.text.strip())
        if amount <= 0:
            raise ValueError("Количество должно быть положительным")

        # Сохраняем количество баллов
        await state.update_data(points_amount=amount)

        await state.set_state(DevState.waiting_for_points_comment)

        comment_msg = await message.answer(
            f"💰 <b>Количество баллов:</b> {amount} G\n\n"
            f"💬 Введите комментарий к начислению:\n\n"
            f"💡 Этот комментарий увидит пользователь",
            reply_markup=dev_cancel_keyboard(),
            parse_mode='HTML'
        )
        await set_last_message_id(state, comment_msg.message_id)

    except ValueError:
        error_msg = await message.answer(
            "❌ Введите корректное положительное число:",
            reply_markup=dev_cancel_keyboard(),
            parse_mode='HTML'
        )
        await set_last_message_id(state, error_msg.message_id)

# 7. Обработка ввода комментария и подтверждение
@router.message(DevState.waiting_for_points_comment)
async def process_points_comment(message: Message, state: FSMContext, bot: Bot):
    """Обработка комментария и запрос подтверждения."""

    await delete_previous_message(state, message.chat.id, bot)
    await message.delete()

    comment = message.text.strip()

    # Сохраняем комментарий
    await state.update_data(points_comment=comment)

    # Получаем все сохраненные данные
    data = await state.get_data()

    confirm_msg = await message.answer(
        f"📋 <b>Проверьте данные перед начислением:</b>\n\n"
        f"👤 Получатель: <b>{data['target_user_name']}</b>\n"
        f"🆔 Username: @{data['target_username'] or 'не указан'}\n"
        f"💰 Сумма: <b>{data['points_amount']}</b> G\n"
        f"💬 Комментарий: <i>{comment}</i>\n\n"
        f"❓ Подтвердить начисление?",
        reply_markup=dev_confirm_keyboard(),
        parse_mode='HTML'
    )
    await set_last_message_id(state, confirm_msg.message_id)

# 8. Подтверждение и выполнение начисления
@router.callback_query(F.data == "dev_confirm_add_points")
async def confirm_add_points(callback: CallbackQuery, state: FSMContext, session: AsyncSession, bot: Bot):
    """Финальное начисление баллов и уведомление пользователя."""

    await callback.answer()

    data = await state.get_data()

    try:
        user_service = UserService(session)

        # Начисляем баллы
        new_balance = await user_service.update_balance(
            telegram_id=data['target_user_id'],
            amount=data['points_amount']
        )

        await state.clear()

        # Уведомляем пользователя
        try:
            await bot.send_message(
                data['target_user_id'],
                f"🎁 <b>Вам начислены баллы!</b>\n\n"
                f"💰 Сумма: <b>+{data['points_amount']}</b> G\n"
                f"💼 Новый баланс: <b>{new_balance}</b> G\n\n"
                f"💬 <b>Комментарий администратора:</b>\n<i>{data['points_comment']}</i>",
                parse_mode='HTML'
            )
        except Exception as e:
            print(f"Не удалось отправить уведомление пользователю: {e}")

        # Уведомляем администратора об успехе
        await delete_previous_message(state, callback.message.chat.id, bot)

        success_msg = await callback.message.answer(
            f"✅ <b>Баллы успешно начислены!</b>\n\n"
            f"👤 Получатель: <b>{data['target_user_name']}</b>\n"
            f"💰 Начислено: <b>{data['points_amount']}</b> G\n"
            f"💼 Новый баланс: <b>{new_balance}</b> G\n\n"
            f"✉️ Пользователь получил уведомление",
            parse_mode='HTML'
        )

        import asyncio
        await asyncio.sleep(3)
        await success_msg.delete()

        await show_dev_menu(callback.message, state, bot)

    except Exception as e:
        await state.clear()
        await callback.message.answer(
            f"❌ Ошибка при начислении баллов: {e}",
            reply_markup=dev_menu_keyboard(),
            parse_mode='HTML'
        )
        await show_dev_menu(callback.message, state, bot)

# --- НОВЫЙ ХЕНДЛЕР: ЭКСПОРТ БАЗЫ ПОЛЬЗОВАТЕЛЕЙ ---

# --- НОВЫЙ ХЕНДЛЕР: ЭКСПОРТ БАЗЫ ПОЛЬЗОВАТЕЛЕЙ ---

# --- НОВЫЙ ХЕНДЛЕР: ЭКСПОРТ БАЗЫ ПОЛЬЗОВАТЕЛЕЙ ---

@router.callback_query(F.data == "dev_export_users")
async def export_users_database(callback: CallbackQuery, session: AsyncSession, bot: Bot, state: FSMContext):
    """Экспорт базы пользователей в Excel."""

    await callback.answer("Формирую базу данных...")

    await delete_previous_message(state, callback.message.chat.id, bot)

    loading_msg = await callback.message.answer(
        "⏳ <b>Экспорт базы пользователей...</b>\n\n"
        "Пожалуйста, подождите. Формируется Excel-файл.",
        parse_mode='HTML'
    )

    try:
        # Получаем всех пользователей из БД
        result = await session.execute(select(User))
        users = result.scalars().all()

        if not users:
            await loading_msg.edit_text(
                "❌ База пользователей пуста.",
                reply_markup=dev_menu_keyboard()
            )
            return

        # Создаем Excel-файл
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Пользователи"

        # Настройка заголовков (БЕЗ "Баланс руб.")
        headers = [
            "UID", "Telegram ID", "Username", "Полное имя",
            "Телефон", "Email", "Дата рождения", "Баланс G",
            "Premium Rate", "Крипто-кошелек",
            "Дата регистрации"
        ]

        # Стиль для заголовков - просто жирный черный текст
        header_font = Font(bold=True, color="000000", size=11)
        header_alignment = Alignment(horizontal="center", vertical="center")

        # Записываем заголовки
        for col_num, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col_num, value=header)
            cell.font = header_font
            cell.alignment = header_alignment

        # Записываем данные пользователей (БЕЗ digital_ruble_balance)
        for row_num, user in enumerate(users, 2):
            ws.cell(row=row_num, column=1, value=user.uid)
            ws.cell(row=row_num, column=2, value=user.telegram_id)
            ws.cell(row=row_num, column=3, value=user.username or "—")
            ws.cell(row=row_num, column=4, value=user.full_name or "—")
            ws.cell(row=row_num, column=5, value=user.phone or "—")
            ws.cell(row=row_num, column=6, value=user.email or "—")
            ws.cell(row=row_num, column=7, value=user.birth_date.strftime("%d.%m.%Y") if user.birth_date else "—")
            ws.cell(row=row_num, column=8, value=float(user.balance_g))
            ws.cell(row=row_num, column=9, value=user.premium_rate or "—")
            ws.cell(row=row_num, column=10, value=user.crypto_wallet or "Не подключен")
            ws.cell(row=row_num, column=11, value=user.registered_at.strftime("%d.%m.%Y %H:%M") if user.registered_at else "—")

        # Автоширина столбцов
        for column in ws.columns:
            max_length = 0
            column_letter = column[0].column_letter
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min(max_length + 2, 50)
            ws.column_dimensions[column_letter].width = adjusted_width

        # Сохраняем файл
        filename = f"users_export_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
        wb.save(filename)

        # Отправляем файл администратору
        await loading_msg.delete()

        file = FSInputFile(filename)
        await callback.message.answer_document(
            document=file,
            caption=f"📊 <b>Экспорт базы пользователей</b>\n\n"
                    f"👥 Всего пользователей: <b>{len(users)}</b>\n"
                    f"📅 Дата экспорта: <b>{datetime.now().strftime('%d.%m.%Y %H:%M')}</b>",
            parse_mode='HTML'
        )

        # Удаляем временный файл
        os.remove(filename)

        await show_dev_menu(callback.message, state, bot)

    except Exception as e:
        await loading_msg.edit_text(
            f"❌ <b>Ошибка при экспорте:</b>\n\n<code>{e}</code>",
            reply_markup=dev_menu_keyboard(),
            parse_mode='HTML'
        )
        import traceback
        traceback.print_exc()

# --- ОСТАЛЬНЫЕ ХЕНДЛЕРЫ ---

# 9. Выход из режима ввода (кнопка Отмена)
@router.callback_query(F.data == "dev_menu_return")
async def exit_input_mode(callback: CallbackQuery, state: FSMContext, bot: Bot):
    """Возвращает в меню разработчика из состояния ввода."""
    await callback.answer("Отменено.")

    # Удаляем сообщение с запросом ввода
    await delete_previous_message(state, callback.message.chat.id, bot)

    # Возвращаемся в меню разработчика
    await show_dev_menu(callback.message, state, bot)

# 10. Сброс FSM-состояний
@router.callback_query(F.data == "dev_clear_fsm")
async def clear_fsm(callback: CallbackQuery, state: FSMContext, bot: Bot):
    """Сбрасывает все FSM-состояния пользователя."""
    await callback.answer("FSM-состояния сброшены.", show_alert=True)
    await state.clear()

    # Возвращаемся в меню разработчика
    await show_dev_menu(callback.message, state, bot)